import time
from openpyxl import Workbook
from types import SimpleNamespace
from openpyxl.styles import PatternFill
from playhouse.shortcuts import model_to_dict
# from rich.table import Table
# from rich.console import Console
# from rich import print


from libraries.findings_models import (
    DailyGenerals,
    DailyFindings,
    DailyMetadata,
    get_unique_agency_numbers,
    get_unique_cog_overs,
    setup_database
)

import libraries.findings_util as findings_util

from libraries.findings_util import (
    op,
    string_to_datetime,
    fetch_from_api,
    today,
    get_query_count,
    rm,
    path_based_on_ext,
    adjust_columns,
    cog_over,
    convert_bools,
)

from libraries.findings_const import (
    FAC_API_BASE
)

import logging
logger = logging.getLogger(__name__)

# https://stackoverflow.com/questions/17755996/how-to-make-a-list-as-the-default-value-for-a-dictionary

# A result is a single award for a single day.

def set_api_key(key):
    findings_util.set_api_key(key)


findings_fields_to_keep = set([
    "report_id",
    "auditee_name",
    "auditee_uei",
    "cog_over",
    "award_reference",
    "reference_number",
    "is_modified_opinion",
    "is_other_matters",
    "is_material_weakness",
    "is_significant_deficiency",
    "is_other_findings",
    "is_questioned_costs",
    "is_repeat_finding",
    "prior_finding_ref_numbers",
])

awards_fields_to_keep = set([
    "report_id",
    "reference_number",
    "award_reference",
    "auditee_name",
    "aln",
    "cog_over",
    "federal_program_name",
    "amount_expended",
    "is_direct",
    "is_major",
    "is_passthrough_award",
    "passthrough_amount",
])

yes_fill = PatternFill(start_color="FFD700",
                       end_color="FFD700", fill_type="solid")


class QParam():
    def __init__(self, date):
        self.date = date


class Result():
    def __init__(self, d):
        self.data = SimpleNamespace(**d)

    def add(self, key, value):
        self.data[key] = value

    def __str__(self):
        return f"{self.data.report_id}"

    def __repr__(self):
        return self.__str__()


class FAC():
    # Takes a list of parameter objects.
    def __init__(self, params):
        # TODO: Remove any dates we have already run and
        # cached locally.
        self.params = params
        self.results = []

    # Fetch the general results.
    # Must start here.
    def general(self, report_id=None):
        po: QParam

        for po in self.params:
            payload = {
                "fac_accepted_date": op("eq", po.date),
                "select": ",".join([
                    "report_id",
                    "auditee_name",
                    "cognizant_agency",
                    "oversight_agency",
                    "auditee_uei",
                ])
            }

            jres = fetch_from_api("general", payload)

            for res in jres:
                if DailyGenerals.select().where(DailyGenerals.report_id == res["report_id"]):
                    logger.debug(f"Skipping {res['report_id']}")
                else:
                    d = {"report_id": res["report_id"],
                         "date": po.date,
                         "auditee_name": res["auditee_name"],
                         "cog_over": cog_over(res["cognizant_agency"], res["oversight_agency"]),
                         "auditee_uei": res["auditee_uei"],
                         }
                    self.results.append(DailyGenerals.create(**d))

    # Now, populate with the findings. This tells us which we need, and
    # which to remove.
    def findings(self, report_id=None):
        print("FINDINGS")
        # console = Console()
        # We should only do things where we have not fetched.
        if report_id:
            gq = DailyGenerals.select().where(DailyGenerals.report_id == report_id)
        else:
            gq = DailyGenerals.select().where(DailyGenerals.findings_count.is_null())
        for dg in gq:
            print(f"\tfindings {dg.report_id}")
            jres = fetch_from_api("findings", {
                "report_id": op("eq", dg.report_id)
            })
            for res in jres:
                ## print(f"res {res}")
                res["cog_over"] = dg.cog_over
                res = res | model_to_dict(dg)
                # We only need a subset of the keys
                # that come back from the API query.
                to_delete = set(res.keys()).difference(findings_fields_to_keep)
                for k in to_delete:
                    del res[k]
                # Make sure booleans are booleans...
                # Peewee does not treat 'N' as False.
                res = convert_bools(res)
                # console.log(res)
                dfq = (DailyFindings
                       .select()
                       .where((DailyFindings.report_id == dg.report_id)
                              & (DailyFindings.award_reference == res["award_reference"])
                              & (DailyFindings.reference_number == res["reference_number"])))
                if dfq.exists():
                    for df in dfq:
                        print(
                            f"\tUpdating {dg.report_id} {res['award_reference']} {res['reference_number']}")
                        (df
                         .update(**res)
                         .where((DailyFindings.report_id == dg.report_id)
                                & (DailyFindings.award_reference == res["award_reference"])
                                & (DailyFindings.reference_number == res["reference_number"]))
                         .execute())
                else:
                    print(
                        f"\tCreating {dg.report_id} {res['award_reference']} {res['reference_number']}")
                    DailyFindings.create(**res)
            dg.date_retrieved = today()
            dg.findings_count = len(jres)
            dg.save()

    def awards(self, report_id=None):
        print("AWARDS")
        # console = Console()

        if report_id:
            gq = DailyGenerals.select().where(DailyGenerals.report_id == report_id)
        else:
            gq = DailyGenerals.select().where(DailyGenerals.awards_count.is_null())
        for dg in gq:
            # For each general
            dfq = (DailyFindings
                   .select()
                   .where(DailyFindings.report_id == dg.report_id))
            awards_count = 0
            # We already have findings loaded.
            # These are the awards that we care about
            for df in dfq:
                # Now, for each row we find, we need to
                # look up more award info.
                jres = fetch_from_api("federal_awards", {
                    "report_id": op("eq", dg.report_id),
                    "award_reference": op("eq", df.award_reference)
                })
                awards_count += 1
                # What comes back are federal awards results
                for res in jres:
                    # Update the appropriate record.
                    res["aln"] = (res["federal_agency_prefix"] +
                           "." + res["federal_award_extension"])
                    # We only need a subset of the keys
                    # that come back from the API query.
                    res = res | model_to_dict(dg)
                    to_delete = set(res.keys()).difference(
                        awards_fields_to_keep)
                    for k in to_delete:
                        del res[k]
                    res = convert_bools(res)
                    # Update the row in question
                    print(f"\tUpdating awards for {df.report_id} {df.award_reference} {df.reference_number}")
                    (df
                     .update(**res)
                     .where((DailyFindings.report_id == dg.report_id)
                            & (DailyFindings.award_reference == df.award_reference)
                            & (DailyFindings.reference_number == df.reference_number))
                     .execute())
        dg.awards_count = awards_count
        dg.save()

    def _add_sheets(self, wb, iter, query):
        # get_unique_agency_numbers()
        for iter_value in iter:
            ws = wb.create_sheet(f"{iter_value}")
            # Put headers on the sheets
            for obj in query(iter_value):
                as_d = model_to_dict(obj)
                ws.append(list(as_d.keys()))
                break
            # Now the values.
            for obj in query(iter_value):
                as_d = model_to_dict(obj)
                ws.append(list(as_d.values()))
            adjust_columns(ws)

    def _cleanup_sheet(self, ws):
        boolean_columns = ["K", "L", "M", "O", "P", "Q", "R", "S", "T", "U"]
        # Trys to go through a sheet and
        # 1. Hyperlink all the report ids,
        # 2. Cleanup all the booleans.
        # The columns are hard-coded to the order
        # they appear from the dump into the sheet.
        try:
            report_ids = []
            for cell in ws["B"]:
                if ("GSAFAC" in cell.value) or ("CENSUS" in cell.value):
                    report_ids.append(cell.value)
                    cell.hyperlink = f"https://app.fac.gov/dissemination/report/pdf/{cell.value}"
                else:
                    pass
            for ndx, cell in enumerate(ws["C"][1:]):
                cell.hyperlink = f"https://app.fac.gov/dissemination/summary/{report_ids[ndx]}"
            for bool_column in boolean_columns:
                for cell in ws[bool_column]:
                    if cell.value == 1:
                        cell.value = "YES"
                    elif cell.value == 0:
                        cell.value = "NO"
                    else:
                        pass
            for bool_column in boolean_columns:
                for cell in ws[bool_column]:
                    if cell.value == "YES":
                        cell.fill = yes_fill
        except:
            pass

    def _remove_default_sheet(self, wb):
        # Try removing the default sheet.
        try:
            del wb['Sheet']
        except:
            pass

    def to_xlsx(self):
        print("TO XLSX")
        wb = Workbook()
        self._add_sheets(
            wb,
            get_unique_agency_numbers(),
            lambda iter_value:
            (DailyFindings
             .select
             ().where(DailyFindings.aln.startswith(iter_value)))
        )
        self._add_sheets(
            wb,
            get_unique_cog_overs(),
            lambda iter_value:
            (DailyFindings
             .select
             ().where(DailyFindings.cog_over == iter_value))
        )

        # Hyperlink the report IDs
        for sheet in wb.worksheets:
            self._cleanup_sheet(sheet)

        self._remove_default_sheet(wb)

        return wb


# @click.command()
# @click.argument('acceptance_date', default="2024-03-02")
# @click.option("--clean", is_flag=True, show_default=True, default=False,)
# @click.option("--omit-generals", is_flag=True, show_default=True, default=False,)
# @click.option("--omit-findings", is_flag=True, show_default=True, default=False,)
# @click.option("--omit-awards", is_flag=True, show_default=True, default=False,)
# @click.option("--report-id", default=None,)
def findings_by_aln(acceptance_date, 
         clean=True, 
         omit_generals=False, 
         omit_findings=False, 
         omit_awards=False, 
         report_id=None):
    acceptance_date = string_to_datetime(acceptance_date)
    db_filename = f"{acceptance_date.strftime('%Y-%m-%d')}.sqlite"
    workbook_filename = f"{acceptance_date.strftime('%Y-%m-%d')}-findings.xlsx"
    # Possibly remove work products
    # If we're only running part of the generation, then
    # do not clean things. That's an error on the user's part.
    if clean and (all(map(lambda v: not v, [omit_generals, omit_findings, omit_awards]))):
        rm(path_based_on_ext(db_filename))

    setup_database(db_filename)

    qparams = []
    qparams.append(QParam(acceptance_date.date()))
    fac = FAC(qparams)

    g0 = g1 = 0
    f0 = f1 = 0
    a0 = a1 = 0

    t0 = time.time()
    if omit_generals:
        print("Skipping general generation")
    else:
        g0 = time.time()
        fac.general(report_id=report_id)
        g1 = time.time()

    if omit_findings:
        print("Skipping findings generation")
    else:
        f0 = time.time()
        fac.findings(report_id=report_id)
        f1 = time.time()

    if omit_awards:
        print("Skipping award generation")
    else:
        a0 = time.time()
        fac.awards(report_id=report_id)
        a1 = time.time()
    t1 = time.time()

    try:
        wb = fac.to_xlsx()
        rm(path_based_on_ext(workbook_filename))
        wb.save(path_based_on_ext(workbook_filename))
        DailyMetadata.create(
            date_retrieved=today(),
            queries_used=get_query_count(),
            time_elapsed=t1-t0,
            time_general=g1-g0,
            time_findings=f1-f0,
            time_awards=a1-a0,
        )
    except:
        print(f"{acceptance_date} NO FINDINGS, NO WORKBOOK")
