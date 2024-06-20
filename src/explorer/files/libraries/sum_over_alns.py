import os
import requests
import sys
import datetime
#from alive_progress import alive_bar
from .aln import ALN
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time

# https://stackoverflow.com/questions/17755996/how-to-make-a-list-as-the-default-value-for-a-dictionary
from collections import defaultdict
from pprint import pprint
from collections import namedtuple as NT

FAC_API_BASE = "https://api.fac.gov"
# This change hard-overrides using the local data.
# This involves leaving out some audits, but it is faster,
# and avoids key limit issues while testing.
# FAC_API_BASE = "http://localhost:3000"
# FAC_API_KEY = os.getenv("API_GOV_KEY")
MAX_RESULTS = 4_000_000
STEP_SIZE = 20000

# Basic headers; intended for use locally as well as remotely.
def BASE_HEADERS(api_key):
    return {
        "X-API-Key": api_key
        }


def load_aln_list(fname):
    alns = set()
    with open(fname, 'r') as fp:
        for line in fp:
            line = line.strip()
            parts = line.split(".")
            if len(parts) == 1:
                alns.add(ALN(parts[0]))
            else:
                alns.add(ALN(parts[0], parts[1]))
    return list(alns)


def op(op, value):
    return f"{op}.{value}"


def string_to_datetime(strdate):
    parts = strdate.split("-")
    return datetime.datetime(int(parts[0]), int(parts[1]), int(parts[2]))


memoize_dates = {}


def get_date(report_id, api_key=""):
    if memoize_dates.get(report_id, False):
        return string_to_datetime(memoize_dates.get(report_id))
    payload = {
        "report_id": op("eq", report_id),
        "select": ",".join(["report_id", "fac_accepted_date"]),
        "api_key": api_key,
    }
    res = requests.get(f"{FAC_API_BASE}/general",
                       params=payload,
                       headers=BASE_HEADERS(api_key))
    jres = res.json()
    if len(jres) == 0:
        print(f"NO DATE FOUND FOR {report_id}")
        sys.exit()
    the_date = jres[0]["fac_accepted_date"]
    memoize_dates[report_id] = the_date
    the_date = string_to_datetime(the_date)
    return the_date


def calculate_for_aln(aln,
                      audit_year="2023",
                      before_acceptance="2023-06-28",
                      api_key=""):
    # What report IDs does this ALN appear in?
    # aln : report_id
    aln_to_report_ids = defaultdict(list)
    # What is the total direct amount on that ALN?
    # aln : total
    aln_to_total = defaultdict(lambda: 0)
    # How many times do we see this ALN?
    # aln : count
    aln_to_count = defaultdict(lambda: 0)
    aln_dates = defaultdict(list)
    before_acceptance = string_to_datetime(before_acceptance)

    # We begin by finding this ALN in the federal_awards table
    payload = {
        "limit": STEP_SIZE - 1,
        "federal_agency_prefix": op("eq", aln.agency),
        "audit_year": op("eq", audit_year),
        "is_direct": op("eq", "Y"),
        "select": ",".join(["report_id", "amount_expended", "is_direct", "federal_agency_prefix", "federal_award_extension"]),
        "api_key": api_key,
    }
    # If they included a program, and not just an agency number...
    if aln.program:
        payload["federal_award_extension"] = op("eq", aln.program)

    url = f"{FAC_API_BASE}/federal_awards"

    for start in range(0, MAX_RESULTS, STEP_SIZE):
        payload["offset"] = start
        
        res = requests.get(url,
                           params=payload)
        jres = res.json()
        len_jres = len(jres)
        # print(f"[{payload['offset']} -> {payload['offset'] + (STEP_SIZE-1)}] Retrieved {len_jres} results...")
        if jres == []:
            break
        elif "code" in jres:
            print("ERROR: ")
            pprint(jres)
            break
        else:
            for r in jres:
                this_date = get_date(r["report_id"], api_key)
                r["fac_accepted_date"] = this_date
                if this_date < before_acceptance:
                    aln_to_report_ids[aln].append(r["report_id"])
                    aln_to_count[aln] = aln_to_count.get(aln, 0) + 1
                    aln_dates[aln].append(this_date)
                    if r["is_direct"] == "Y":
                        aln_to_total[aln] = aln_to_total.get(
                            aln, 0) + r["amount_expended"]
            if len_jres < STEP_SIZE:
                break

    # return (str(aln), aln_to_report_ids, aln_to_total, aln_to_count)
    return Results(audit_year, str(aln), aln_to_report_ids[aln], aln_to_total[aln], aln_to_count[aln])


def fac_weight_fun(reports, awards, dollars):
    v = (0.485 * reports) + (0.485 * awards) + (0.03 * dollars)
    return round(v, 3)


class Results():
    def __init__(self, audit_year, aln, report_ids, total_dollars, award_count):
        self.audit_year = audit_year
        self.aln = aln
        self.report_ids = set(report_ids)
        self.total_dollars = total_dollars
        self.award_count = award_count

    def __str__(self):
        return f"{self.aln} rids: {len(self.report_ids)} $: {self.total_dollars} awards: {self.award_count}"

    def __repr__(self):
        return self.__str__()

    def to_csv(self):
        return f"{self.audit_year},{self.aln},{len(self.report_ids)},{self.award_count},{self.total_dollars}"


class ResultSummary():
    def __init__(self, agency_number):
        self.agency_number = agency_number
        self.results = defaultdict(list)
        self.alns = defaultdict(list)
        self.report_counts = defaultdict(list)
        self.award_counts = defaultdict(list)
        self.total_dollars = defaultdict(list)
        self.pct_of_reports = defaultdict(list)
        self.pct_of_awards = defaultdict(list)
        self.pct_of_dollars = defaultdict(list)
        self.fac_weights = defaultdict(list)

    def add_result(self, audit_year, r):
        self.results[audit_year].append(r)

    def prep_report(self):
        for ay, rs in self.results.items():
            for r in rs:
                self.alns[ay].append(r.aln)
                self.report_counts[ay].append(len(r.report_ids))
                self.award_counts[ay].append(r.award_count)
                self.total_dollars[ay].append(int(r.total_dollars))
            try:
                self.pct_of_reports[ay] = list(
                    map(lambda n: round(n / sum(self.report_counts[ay])*100, 3), self.report_counts[ay]))
                self.pct_of_awards[ay] = list(
                    map(lambda n: round(n / sum(self.award_counts[ay])*100, 3), self.award_counts[ay]))
                self.pct_of_dollars[ay] = list(
                    map(lambda n: round(n / sum(self.total_dollars[ay])*100, 3), self.total_dollars[ay]))
                self.fac_weights[ay] = list(map(fac_weight_fun,
                                                self.pct_of_reports[ay],
                                                self.pct_of_awards[ay],
                                                self.pct_of_dollars[ay]))
            except:
                print("REPORT COUNTS", self.report_counts)
                print("AWARD COUNTS", self.award_counts)
                print("TOTAL DOLLARS", self.total_dollars)
                
    def report_as_xlsx(self):
        self.prep_report()
        wb = Workbook()
        ws = wb.create_sheet("Overview")
        df = pd.DataFrame({
            "note": [
                "All values rounded to 3 places.", 
                "FAC weight is (0.485 * pct_rpt) + (0.485 * pct_awd) + (0.03 * pct_$)",
                "FAC weight can be used for estimating opdiv contribution, if desired.",
                ]
        })
        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)

        for ay, _ in self.results.items():
            ws = wb.create_sheet(f"AY{ay}")
            df = pd.DataFrame(
                {
                    "aln": self.alns[ay],
                    "report_count": self.report_counts[ay],
                    "award_count": self.award_counts[ay],
                    "total_dollars": self.total_dollars[ay],
                    "pct_of_reports": self.pct_of_reports[ay],
                    "pct_of_awards": self.pct_of_awards[ay],
                    "pct_of_dollars": self.pct_of_dollars[ay],
                    "fac_weight": self.fac_weights[ay]
                }
            )
            for r in dataframe_to_rows(df, index=True, header=True):
                ws.append(r)
        del wb['Sheet']
        wb.save(f"agency-{self.agency_number}-distribution.xlsx")


def get_alns_by_agency_number(audit_year, agency_number, api_key="NO_API_KEY"):
    payload = {
        "federal_agency_prefix": op("eq", agency_number),
        "select": "federal_award_extension",
        "audit_year": op("eq", audit_year),
    }
    url = f"{FAC_API_BASE}/federal_awards"
    all_alns = set()
    
    for start in range(0, MAX_RESULTS, STEP_SIZE):
        payload = payload | {
            "offset": start,
            "api_key": api_key,
        }
        res = requests.get(url,
                           params=payload,
                          )
        jres = res.json()
        if jres == []:
            break
        elif "code" in jres:
            print("ERROR: ")
            pprint(jres)
            break
        else:
            # Don't bother with another call if we had fewer than the max.
            for r in jres:
                all_alns.add(ALN(agency_number, r["federal_award_extension"]))

    return all_alns

# @click.command()
# @click.argument('list_of_alns')
# @click.option('--audit-years', default="2023", help='Audit year')
# @click.option('--before-acceptance', default="2023-06-28", help="Acceptance date")
# @click.option("--distinct-alns-for-agency", default=None, help="Each distinct aln under an agency number.")
def sum_over_alns(list_of_alns, audit_years, before_acceptance, distinct_alns_for_agency, api_key):
    RS = ResultSummary(distinct_alns_for_agency)

    for audit_year in list(map(lambda y: int(y), audit_years.split(","))):
        if distinct_alns_for_agency:
            alns = get_alns_by_agency_number(
                audit_year, 
                distinct_alns_for_agency,
                api_key
                )
        else:
            alns = load_aln_list(list_of_alns)
        for ndx, aln in enumerate(sorted(alns, key=lambda a: f"{a.agency}.{a.program}")):
            print(f"Calculating for {aln} [{ndx + 1} of {len(alns)}]") 
            result = calculate_for_aln(aln,
                                       audit_year=audit_year,
                                       before_acceptance=before_acceptance,
                                       api_key=api_key,
                                      )
            print(result)
            RS.add_result(audit_year, result)

    RS.report_as_xlsx()
